"""Generate an Excel (.xlsx) version of the MASTER_SUMMARY §0 all-SKU matrix.

Pulls per-interaction values straight from build_summaries.derive() (so the sheet matches
the §0 table, with real numeric cells). Sheet 1 = the SKU matrix (agent names hyperlinked to
their GitHub-hosted docs); Sheet 2 = the column legend + billing-alignment notes.

Usage: python scripts/master_table_xlsx.py [out.xlsx]
"""
import importlib.util
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
GITHUB_BASE = "https://github.com/jasmeetsb/agent-cost-estimator/blob/main"

# Load build_summaries to reuse derive(), PACKAGES, LINKS (single source of truth).
_spec = importlib.util.spec_from_file_location("bs", REPO / "scripts" / "build_summaries.py")
bs = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(bs)

# memory_assistant legacy row (mirrors build_summaries.master()); lacks the newer fields.
MA = {"title": "memory_assistant", "in_tok": 3398, "out_tok": 1605, "calls": 5.75,
      "vcpu_sec": 39.0, "gib_sec": 560.0, "sess": 11.5, "gen_tok": 2493,
      "mem_retrieved": 2.5, "images": 0, "c_total": 0.0165}
_ARCHE = ["conversational_chatbot", "workflow_operator", "autonomous_researcher", "multi_agent_orchestrator"]
_USECASE = ["financial_advisor", "academic_research", "marketing_agency", "blogger_agent"]


def _sortk(r):
    pkg = r.get("pkg", r.get("title", ""))
    if pkg in _ARCHE:
        return (0, _ARCHE.index(pkg))
    if pkg in _USECASE:
        return (1, _USECASE.index(pkg))
    return (2, -r.get("in_tok", 0))


# (header, derive-key-or-fn, number_format)
COLS = [
    ("Agent", lambda d: d["title"], "@"),
    ("Interactions", lambda d: d.get("n"), "0"),
    ("Total turns", lambda d: d.get("total_turns") or None, "0"),
    ("Input tokens", lambda d: d["in_tok"], "#,##0"),
    ("Output tokens", lambda d: d["out_tok"], "#,##0"),
    ("Model calls", lambda d: d["calls"], "0.0"),
    ("vCPU-seconds", lambda d: d["vcpu_sec"], "0.0"),
    ("GiB-seconds", lambda d: d["gib_sec"], "0.0"),
    ("Session events", lambda d: d["sess"], "0.0"),
    ("Mem-gen tokens", lambda d: d["gen_tok"], "#,##0"),
    ("Memories retrieved", lambda d: d["mem_retrieved"], "0.00"),
    ("Firestore writes", lambda d: d.get("fs_writes_pi", 0), "0.00"),
    ("Firestore reads", lambda d: d.get("fs_reads_pi", 0), "0.00"),
    ("RAG queries", lambda d: d.get("rag_pi", 0), "0.00"),
    ("Web grounding", lambda d: d.get("web_ground_pi", d.get("web_searches", 0)), "0.00"),
    ("Imagen images", lambda d: d.get("images", 0), "0"),
    ("$/interaction", lambda d: d["c_total"], "$0.0000"),
]

LEGEND = [
    ("Column", "Meaning (per interaction, avg over Interactions unless noted)"),
    ("Agent", "Agent architecture (links to its GitHub-hosted summary)."),
    ("Interactions", "Number of interactions tested (sample size for every average in the row)."),
    ("Total turns", "Total user turns across the experiment (Σ turns over all interactions)."),
    ("Input / Output tokens", "Gemini prompt (incl. cached) / output (candidates + thinking) tokens. BILLING-ACCURATE."),
    ("Model calls", "Model invocations per interaction. NOT a billing unit (Gemini bills tokens) — a usage driver."),
    ("vCPU-s / GiB-s", "Agent Runtime allocation-time, amortized over the window incl. idle. ESTIMATE, not actual instance-hours."),
    ("Session events", "Events appended to Sessions. Observed (not metered); excludes session storage GiB-hr."),
    ("Mem-gen tokens", "Memory Bank generation tokens (add_session_to_memory). Priced at input rate (proxy); excludes monthly storage."),
    ("Memories retrieved", "Memory Bank memories returned via load_memory. BILLING-ACCURATE ($0.5/1K)."),
    ("Firestore writes/reads", "Firestore document ops (save_note/load_note). BILLING-ACCURATE. NOTE: Firestore is not in the GE AP calculator (added as a representative operational DB)."),
    ("RAG queries", "Vertex AI Search queries. BILLING-ACCURATE ($1.5/1K); indexed-data storage not captured."),
    ("Web grounding", "Google Search grounded query-turns (web-research AgentTool calls). PROXY/lower bound; billed per grounded prompt ($14/1K)."),
    ("Imagen images", "Images generated (Imagen / gemini-2.5-flash-image). BILLING-ACCURATE (flat-rate estimate)."),
    ("$/interaction", "Derived cost = Σ(usage × catalog LIST price); incl. Model Armor. REFERENCE ONLY — not billed dollars (no discounts/CUDs)."),
    ("", ""),
    ("Uncaptured SKUs", "Apigee, BigQuery, Veo, Maps grounding, Agent Sandbox, Agent Gateway, Cloud Logging/Trace/Monitoring, Agent Evaluation — parked/deferred/pending."),
]


def main():
    out = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else REPO / "agent_summaries" / "MASTER_SUMMARY.xlsx"
    rows = [bs.derive(p) for p in bs.PACKAGES] + [MA]
    rows.sort(key=_sortk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SKU matrix"
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    for c, (name, _, _) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ri, d in enumerate(rows, 2):
        for ci, (name, fn, fmt) in enumerate(COLS, 1):
            val = fn(d)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.number_format = fmt
            if ci == 1:
                link = bs.LINKS.get(d["title"])
                if link:
                    cell.hyperlink = f"{GITHUB_BASE}/agent_summaries/{link}"
                    cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "B2"  # freeze header row + agent column
    ws.column_dimensions["A"].width = 34
    for c in range(2, len(COLS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows) + 1}"

    wl = wb.create_sheet("Legend")
    for ri, (a, b) in enumerate(LEGEND, 1):
        ca = wl.cell(row=ri, column=1, value=a); cb = wl.cell(row=ri, column=2, value=b)
        if ri == 1:
            ca.font = cb.font = Font(bold=True)
        else:
            ca.font = Font(bold=True)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
    wl.column_dimensions["A"].width = 24
    wl.column_dimensions["B"].width = 100

    wb.save(out)
    print(f"wrote {out}  ({len(rows)} agents, {len(COLS)} columns)")


if __name__ == "__main__":
    main()
